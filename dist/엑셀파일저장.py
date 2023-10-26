import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Create a new workbook and select the active worksheet
workbook = Workbook()
sheet = workbook.active

# Define the header row
header = ["ID", "Name", "Quantity", "Price"]

# Write the header row to the worksheet
for col_idx, header_text in enumerate(header, 1):
    col_letter = get_column_letter(col_idx)
    sheet[f"{col_letter}1"] = header_text

# Generate 100 rows of sample data (you can replace this with your own data generation logic)
for row_idx in range(2, 102):
    # Example data: ID, Name, Quantity, Price
    product_id = row_idx - 1
    product_name = f"Product {product_id}"
    quantity = 10  # Example quantity
    price = 500  # Example price

    # Write data to the worksheet
    sheet[f"A{row_idx}"] = product_id
    sheet[f"B{row_idx}"] = product_name
    sheet[f"C{row_idx}"] = quantity
    sheet[f"D{row_idx}"] = price

# Save the workbook to the specified path
file_path = r"C:\work\sales.xlsx"
workbook.save(file_path)

print(f"Sales list has been created and saved at: {file_path}")
